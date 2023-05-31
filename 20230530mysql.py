from sqlalchemy import create_engine
from openpyxl import load_workbook,Workbook
import pandas as pd
import numpy as np
import warnings,os,sys,datetime
import pymysql

#屏蔽编译告警输出
warnings.filterwarnings("ignore")

#判断当前文件夹加内是否存在文件“掉话统计.xlsx”
if not os.path.exists('掉话统计.xlsx'):
    print('当前文件夹不存在文件“掉话统计.xlsx”！程序将自动创建。')
    wb = Workbook()
    wb.save('掉话统计.xlsx')
    wb.close()
    del wb
else:
    print('当前文件夹存在文件“掉话统计.xlsx”！程序将自动清空数据！')
    wb = load_workbook('掉话统计.xlsx')
    ws = wb.active
    ws.title = 'Temp'   # 将活动工作表重命名为'Temp'，以便日后删除
    
    for sheet in wb.worksheets:
        if sheet != ws:
            wb.remove(sheet)
    #新建一个工作表
    wb.create_sheet('Sheet1')
    #删除活动工作表
    wb.remove(wb['Temp'])
    
    wb.save('掉话统计.xlsx')
    wb.close()
    del wb
#定义函数，判断用户输入的格式是否正确
def is_valid_date(date_text):
    try:
        datetime.datetime.strptime(date_text, '%Y%m%d')
        return True
    except ValueError:
        return False
#--------------------------读取CHR数据-------------------------
HOST = '10.169.254.90'
PORT = '3306'
USER = 'root'
PASSWORD = 'Wbzlrx760430!'
DB = 'chr'

engine = create_engine('mysql+pymysql://%s:%s@%s:%s/%s?charset=utf8' % (USER, PASSWORD, HOST, PORT, DB))

tables = pd.read_sql("show tables like 'chr%%';",engine)

print(tables)

#获取所有行的第一个元素，即不包含名称
tables = tables.iloc[:,0]

print(tables)


#判断输入日期是否正确，开始日期早于结束日期

print('程序运行过程中所使用的日期为8位半角数字，格式为“YYYYMMDD”。')
print('即4位年份数字2位月份数字和2位日期数字。\n')

date_begin = input('请输入统计日期区间的开始日期：')
date_end = input('\n请输入统计日期区间的结束日期：')

print(date_begin)
print(date_end)
